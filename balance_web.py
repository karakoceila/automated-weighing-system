import re
import time
import threading
from datetime import datetime
import os

import serial
from openpyxl import Workbook, load_workbook
from flask import Flask, redirect, render_template_string, url_for, Response, request
import winsound

# ================== CONFIG ==================
PORT_1 = "COM3"
BAUD = 9600
EXCEL_FILE = "poids_caisses.xlsx"

# Plage poids acceptée (kg)
POIDS_MIN = 13.080
POIDS_MAX = 13.220

# “Stabilité avec vent”
STABLE_SECONDS = 3.0      # durée d'observation
TOL_STABILITE = 0.005     # tolérance vent (ex: 0.005 = 5 g)

# Logique caisse
SEUIL_VIDE_KG = 0.20      # pour considérer que la caisse est retirée
# ===========================================

app = Flask(__name__)

weight_re = re.compile(r"([+-]?\d+(?:[.,]\d+)?)\s*kg", re.IGNORECASE)

# Données partagées par balance
data_scales = {
    1: {
        'history': [],
        'session_start': datetime.now(),
        'current_weight': None,
        'com_ok': False,
        'status_text': 'Démarrage...',
        'etat': 'ATTENTE_CAISSE',
        'samples': [],
        'lock': threading.Lock()
    }
}

# verrou global pour les accès Excel
wb_lock = threading.Lock()


def get_or_create_wb(path):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws = wb.active
        ws.title = "Caisses"
        # colonne "Balance" ajoutée pour différencier les balances
        ws.append(["Date/Heure", "Balance", "Poids caisse (kg)", "Port", "Plage", "Note"])
    return wb, ws


def parse_weight(line: str):
    # on garde "ST..." uniquement (comme ton ancienne version stable)
    if not line.strip().upper().startswith("ST"):
        return None
    m = weight_re.search(line)
    if not m:
        return None
    val = m.group(1).replace(",", ".")
    try:
        return float(val)
    except ValueError:
        return None


def in_range(w: float) -> bool:
    return POIDS_MIN <= w <= POIDS_MAX


def window_ready(now: float, samples_list) -> bool:
    # On garde seulement les samples des STABLE_SECONDS dernières secondes
    cutoff = now - STABLE_SECONDS
    samples_list[:] = [(t, w) for (t, w) in samples_list if t >= cutoff]
    # fenêtre ok si on couvre presque toute la durée
    return len(samples_list) >= 6 and (samples_list[-1][0] - samples_list[0][0]) >= (STABLE_SECONDS - 0.2)


def window_stable(samples_list) -> bool:
    # Stable si amplitude <= TOL_STABILITE
    ws = [w for (_, w) in samples_list]
    if not ws:
        return False
    return (max(ws) - min(ws)) <= TOL_STABILITE


def loop_pesage(scale_id: int):
    """Boucle de lecture pour une balance donnée (1 ou 2)."""
    ds = data_scales[scale_id]

    # choisir le port (seule la balance 1 est utilisée)
    port = PORT_1

    wb, ws = get_or_create_wb(EXCEL_FILE)
    while True:
        try:
            ser = serial.Serial(port, BAUD, bytesize=8, parity="N", stopbits=1, timeout=1)
            with ds['lock']:
                ds['com_ok'] = True
                ds['status_text'] = "COM OK - Prêt"
            print(f"Connecté sur {port} (Balance {scale_id})")
        except Exception as e:
            with ds['lock']:
                ds['com_ok'] = False
                ds['status_text'] = "COM OFF - vérifier câble/port"
            print(f"Erreur ouverture port série {port} (Balance {scale_id}):", e)
            time.sleep(2)
            continue

        # lecture boucle
        while True:
            try:
                raw = ser.readline()
            except Exception as e:
                print("Erreur lecture série:", e)
                with ds['lock']:
                    ds['com_ok'] = False
                    ds['status_text'] = "COM OFF - déconnecté"
                try:
                    ser.close()
                except:
                    pass
                time.sleep(1)
                break  # on retente la connexion

            if not raw:
                continue

            line = raw.decode(errors="ignore").strip()
            w = parse_weight(line)
            if w is None:
                continue

            now = time.time()

            with ds['lock']:
                ds['current_weight'] = w

            # Gestion états caisse (par balance)
            if ds['etat'] == "ATTENTE_CAISSE":
                # on remplit la fenêtre temporelle
                ds['samples'].append((now, w))
                if not window_ready(now, ds['samples']):
                    with ds['lock']:
                        ds['status_text'] = "AJUSTER - attente stabilité 3s"
                    continue

                if not window_stable(ds['samples']):
                    with ds['lock']:
                        ds['status_text'] = "AJUSTER - vent/mouvements"
                    continue

                # Ici : stable “avec vent” sur 3s
                w_min = min(x[1] for x in ds['samples'])
                w_max = max(x[1] for x in ds['samples'])   # on prend la plus haute
                w_save = round(w_max, 3)

                if in_range(w_save):
                    ts = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

                    # Excel (synchronisé)
                    with wb_lock:
                        # insérer sous l'entête pour garder les plus récents en haut
                        ws.insert_rows(2)
                        ws.cell(row=2, column=1, value=ts)
                        ws.cell(row=2, column=2, value=f"Balance {scale_id}")
                        ws.cell(row=2, column=3, value=w_save)
                        ws.cell(row=2, column=4, value=port)
                        ws.cell(row=2, column=5, value=f"{POIDS_MIN:.3f}-{POIDS_MAX:.3f}")
                        ws.cell(row=2, column=6, value="OK")
                        wb.save(EXCEL_FILE)

                    with ds['lock']:
                        ds['history'].append({"ts": ts, "weight": w_save})
                        ds['status_text'] = "OK - enregistrée ✅"
                        ds['etat'] = "ATTENTE_VIDE"
                        ds['samples'] = []

                    try:
                        winsound.Beep(1000, 500)
                    except Exception:
                        pass
                else:
                    with ds['lock']:
                        ds['status_text'] = f"Hors plage ({w_save:.3f} kg)"

            elif ds['etat'] == "ATTENTE_VIDE":
                # attendre que la caisse soit retirée (poids proche de 0)
                ds['samples'].append((now, w))
                # ne garder que dernières 1s
                cutoff = now - 1.0
                ds['samples'] = [(t, ww) for (t, ww) in ds['samples'] if t >= cutoff]
                vals = [ww for (_, ww) in ds['samples']]
                if vals and max(vals) < SEUIL_VIDE_KG:
                    with ds['lock']:
                        ds['etat'] = "ATTENTE_CAISSE"
                        ds['status_text'] = "Prêt - déposer caisse"
                        ds['samples'] = []


HTML_TEMPLATE = """
<!doctype html>
<html lang="fr">
<head>
    <meta charset="utf-8">
    <title>Balance de pesée Tifra Fish</title>
    <meta http-equiv="refresh" content="2">
    <style>
        body { font-family: Arial, sans-serif; background:#020b1a; color:#f9fafb; margin:0; }
        .header { display:flex; align-items:center; justify-content:space-between;
                            padding:10px 30px; background:#010713; }
        .logo img { height:160px; }
        .title { text-align:center; flex:1; }
        .title h1 { margin:0; font-size:34px; }
        .title h2 { margin:4px 0 0; font-size:18px; font-weight:normal; color:#9ca3af; }
        .title a { color:#9ca3af; text-decoration:none; transition: color 0.15s ease, transform 0.12s ease; display:inline-flex; align-items:center; }
        .title a:hover { color:#ffffff; text-decoration:underline; transform:translateY(-2px); }
        .content { text-align:center; padding:25px; }
        .big { font-size:52px; margin: 10px 0; }
        .mid { font-size:22px; margin:5px 0; }
        .weight-box {
                margin: 25px auto 10px;
                padding: 25px 40px;
                border: 4px solid #00FF00;
                border-radius: 30px;
                width: 70%;
                max-width: 1000px;
                background:#020617;
        }
        .big-weight {
        font-size:130px;
        letter-spacing:2px;
        color:#32FF00;                 /* Un vert plus doux */
        text-shadow:0 0 8px #1aff00;   /* Glow plus léger */
        }
        table { margin:auto; border-collapse:collapse; margin-top:20px; min-width:650px; }
        th, td { border:1px solid #374151; padding:10px 18px; font-size:22px; }
        th { background:#0b1120; }
        td { color:#e5e7eb; }
        .btn {
                margin-top:20px; padding:10px 24px; font-size:20px;
                border:none; border-radius:999px;
                background:#22c55e; color:white; cursor:pointer;
        }
        .btn:hover { background:#16a34a; }
        .btn-secondary {
                margin-top:10px; padding:8px 20px; font-size:18px;
                border:none; border-radius:999px;
                background:#0ea5e9; color:white; cursor:pointer;
        }
        .btn-secondary:hover { background:#0284c7; }
        .footer { margin-top:25px; font-size:14px; color:#9ca3af; text-align:center; padding-bottom:10px; }
        .last-weights td { color:#00FF00; }
        .led { width:18px; height:18px; border-radius:50%; display:inline-block; margin-right:8px; vertical-align:middle; }
        .led.green { background:#0a0; box-shadow:0 0 8px #1aff00; }
        .led.red { background:#c00; box-shadow:0 0 6px #ff6b6b; animation: blink 1s infinite; }
        @keyframes blink { 0% { opacity:1; transform:scale(1); } 50% { opacity:0.18; transform:scale(0.92); } 100% { opacity:1; transform:scale(1); } }
        .status-pill { padding:6px 12px; border-radius:12px; background:#07202a; color:#9ca3af; display:inline-block; }
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">
            <img src="{{ url_for('static', filename='tifra_fish.png') }}" alt="Tifra Fish">
        </div>
        
        <div class="title">
            <h1>BALANCE DE PESÉE AUTOMATISÉE</h1>
            <h2>Client : Tifra Fish &nbsp;|&nbsp; Fournisseur : AgynTech Security</h2>
            <div style="margin-top:8px; display:flex; justify-content:center; gap:120px;">
                <a href="/balance/1" style="display:flex;align-items:center;gap:8px;color:#32FF00;text-decoration:none;">
                    <span class="led {% if com_ok_1 %}green{% else %}red{% endif %}"></span>
                    <span style="font-size:14px;">Balance 1</span>
                </a>
            </div>
        </div>
        <div class="logo">
            <img src="{{ url_for('static', filename='agyntech_logo.png') }}" alt="AgynTech">
        </div>
    </div>

    <div class="content">
        <div class="mid">Session démarrée le : {{ session_start }}</div>
        <div class="big">Nombre de caisses : {{ total }}</div>

        <div class="weight-box" style="border-color: {{ color_border }};">
            {% if current_weight is not none %}
                <div class="big-weight" style="color: {{ color_weight }}; text-shadow:0 0 8px {{ color_weight }};">{{ "%.2f"|format(current_weight) }} kg</div>
            {% else %}
                <div class="big-weight" style="color: {{ color_weight }}; text-shadow:0 0 8px {{ color_weight }};">0.00 kg</div>
            {% endif %}
        </div>

        {% if last %}
        <h2>Dernières caisses validées</h2>
        <table class="last-weights">
            <tr><th>#</th><th>Date / heure</th><th>Poids (kg)</th></tr>
            {% for i, item in last %}
                <tr>
                    <td>{{ i }}</td>
                    <td>{{ item.ts }}</td>
                    <td>{{ "%.3f"|format(item.weight) }}</td>
                </tr>
            {% endfor %}
        </table>
        {% else %}
            <p style="font-size:20px;margin-top:30px;">Aucune caisse encore pesée.</p>
        {% endif %}

        <form method="post" action="/reset/{{ sid }}">
            <button class="btn" type="submit">Commencer / Réinitialiser la session</button>
        </form>
        <form method="get" action="/csv/{{ sid }}">
            <button class="btn-secondary" type="submit">Télécharger le rapport (CSV)</button>
        </form>

        
    </div>

    <div class="footer">
        📞 Support AgynTech : 0554752037 &nbsp;|&nbsp; Page locale actualisée toutes les 2 secondes.
    </div>
</body>
</html>
"""


    # COMBINED_HTML removed — root now shows a selector to pick a balance


@app.route("/")
def root_index():
    return redirect(url_for('balance', sid=1))


@app.route('/balance/<int:sid>', methods=['GET'])
def balance(sid: int):
    if sid not in data_scales:
        return "Balance inconnue", 404
    ds = data_scales[sid]
    with ds['lock']:
        total = len(ds['history'])
        # afficher seulement les 2 dernières caisses, triées du plus récent au plus ancien
        last_raw = ds['history'][-2:]
        last_items = list(enumerate(reversed(last_raw), start=1))
        # choisir couleurs selon la balance (ici, seule la balance 1 existe)
        color_border = '#32FF00'
        color_weight = '#32FF00'
        # préparer état LED global pour l'en-tête
        com_ok_1 = data_scales[1]['com_ok']
        return render_template_string(
            HTML_TEMPLATE,
            total=total,
            last=last_items,
            session_start=ds['session_start'].strftime("%d-%m-%Y %H:%M:%S"),
            current_weight=ds['current_weight'],
            com_ok=ds['com_ok'],
            com_ok_1=com_ok_1,
            status_text=ds['status_text'],
            sid=sid,
            color_border=color_border,
            color_weight=color_weight,
        )


@app.route('/reset/<int:sid>', methods=['POST'])
def reset_poste(sid: int):
    if sid not in data_scales:
        return "Balance inconnue", 404
    ds = data_scales[sid]
    with ds['lock']:
        ds['history'] = []
        ds['session_start'] = datetime.now()
    return redirect(url_for('balance', sid=sid))


@app.route('/csv/<int:sid>', methods=['GET'])
def csv_export(sid: int):
    if sid not in data_scales:
        return "Balance inconnue", 404
    import csv
    from io import StringIO

    ds = data_scales[sid]
    si = StringIO()
    writer = csv.writer(si, delimiter=';')
    writer.writerow(["Date/Heure", "Poids caisse (kg)"])
    with ds['lock']:
        for item in ds['history']:
            writer.writerow([item["ts"], f"{item['weight']:.3f}"])

    output = si.getvalue()
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=rapport_pesee_balance_{sid}.csv"}
    )


@app.route('/status/<int:sid>')
def status(sid: int):
    if sid not in data_scales:
        return {}, 404
    ds = data_scales[sid]
    with ds['lock']:
        w = None if ds['current_weight'] is None else round(ds['current_weight'], 3)
        return {
            'com_ok': ds['com_ok'],
            'weight': w,
            'status': ds['status_text'],
            'history': ds['history'][-10:]
        }


if __name__ == "__main__":
    # Support running a single-scale instance via env var SINGLE_SCALE ("1" or "2")
    single = os.environ.get("SINGLE_SCALE")
    flask_port = int(os.environ.get("FLASK_PORT", "5000"))
    if single in ("1", "2"):
        sid = int(single)
        t = threading.Thread(target=loop_pesage, args=(sid,), daemon=True)
        t.start()
        print(f"Démarrage mode single-scale: Balance {sid} on port {flask_port}")
        app.run(host="0.0.0.0", port=flask_port, debug=False)
    else:
        # default: start only balance 1
        t1 = threading.Thread(target=loop_pesage, args=(1,), daemon=True)
        t1.start()
        print("Démarrage mode single: Balance 1 on port 5000")
        app.run(host="0.0.0.0", port=5000, debug=False)
