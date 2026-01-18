import re
import time
from datetime import datetime
import os

import serial
from openpyxl import Workbook, load_workbook
import winsound

# CONFIG
PORT = "COM3" # ⚠️ change si besoin (COM4, COM5...)
BAUD = 9600
EXCEL_FILE = "poids_lp7510.xlsx"

MIN_CAISSE_KG = 0.5
SEUIL_VIDE_KG = 0.2
NB_ECHANTILLONS = 8
TOL_STABILITE = 0.02

weight_re = re.compile(r"([+-]?\d+(?:[.,]\d+)?)\s*kg", re.IGNORECASE)


def get_or_create_wb(path):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Caisses"
        ws.append(["Date/Heure", "Poids caisse (kg)", "Trame brute"])
    return wb, ws


def parse_line(line: str):
    if not line.strip().upper().startswith("ST"):
        return None
    m = weight_re.search(line)
    if not m:
        return None
    val = m.group(1).replace(",", ".")
    try:
        return float(val)
    except ValueError:
        return None


def main():
    wb, ws = get_or_create_wb(EXCEL_FILE)

    try:
        ser = serial.Serial(PORT, BAUD, bytesize=8, parity="N", stopbits=1, timeout=1)
    except Exception as e:
        print("Erreur port série :", e)
        return

    print("Connecté sur", PORT)
    print("Attente de caisses...")

    etat = "ATTENTE_CAISSE"
    fenetre = []

    while True:
        raw = ser.readline()
        if not raw:
            continue

        try:
            line = raw.decode(errors="ignore").strip()
        except:
            continue

        w = parse_line(line)
        if w is None:
            continue

        fenetre.append(w)
        if len(fenetre) > NB_ECHANTILLONS:
            fenetre.pop(0)

        if len(fenetre) < NB_ECHANTILLONS:
            continue

        w_min = min(fenetre)
        w_max = max(fenetre)
        stable = (w_max - w_min) <= TOL_STABILITE

        if etat == "ATTENTE_CAISSE":
            if stable and w_max >= MIN_CAISSE_KG:
                poids_caisse = round(sum(fenetre) / len(fenetre), 3)
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                ws.append([ts, poids_caisse, line])
                wb.save(EXCEL_FILE)

                print(f"[OK] {ts} => {poids_caisse} kg")

                try:
                    winsound.Beep(1000, 300)
                except:
                    pass

                etat = "ATTENTE_VIDE"

        else: # ATTENTE_VIDE
            if stable and w_max <= SEUIL_VIDE_KG:
                print("Balance vide. Prêt pour la caisse suivante.")
                etat = "ATTENTE_CAISSE"


if __name__ == "__main__":
    main()